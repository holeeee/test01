import sys
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import pyautogui
from time import sleep
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook

# 29
# 비로그인 상태로 회의실 만들기 팝업창이 뜨는지 확인
def test29():
    btn_elm = driver.find_element_by_css_selector('#root > div > div._227SmuIIA_SGYBpxZc55bj > div._11HDw8Md_fZm5HmzzPFclC > button')
    btn_elm.click()
    sleep(2)
    try:
        elm = driver.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[2]/div/p[1]/span[contains(text(), "구매 문의")]')
        ws['B5'] = '29'
        ws['C5'] = 'PASS'
        btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[1]/button')
        btn_elm.click()
        return 1
    except:
        ws['B5'] = '29'
        ws['C5'] = 'FAIL'
        btn_elm = driver.find_element_by_xpath('/html/body/div/div/div/div[1]/button')
        btn_elm.click()
        return 0

# 11
# Tmax Admin계정으로 회의실 만들기 팝업창이 뜨지 않는지 확인
def test11():
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > button')
    btn_elm.click()
    sleep(2)
    elm = driver.find_element_by_id('username')
    elm.clear()
    elm.send_keys('yoojung_jang@tmax.co.kr')
    elm = driver.find_element_by_id('password').send_keys('qwer1234!')
    btn_elm = driver.find_element_by_id('kc-form-buttons').click()
    sleep(2)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._227SmuIIA_SGYBpxZc55bj > div._11HDw8Md_fZm5HmzzPFclC > button')
    btn_elm.click()
    sleep(2)
    try:
        elm = driver.find_element_by_id('input_room_name')
        ws['B6'] = '11'
        ws['C6'] = 'PASS'
        btn_elm = driver.find_element_by_xpath('/html/body/div/div/div/div[1]/button')
        btn_elm.click()
        return 1
    except:
        ws['B6'] = '11'
        ws['C6'] = 'FAIL'
        btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[1]/button')
        btn_elm.click()
        return 0

# 10
# Tmax Admin계정으로 비밀번호 변경 후 로그인 되는지 확인
def test10():
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > a > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div.ICjgBEGgF_4nnT6DOdQ9t > div > dl:nth-child(2) > dt > a > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div.RH4-LqVZJjx9RYchbxlwH > table > tr:nth-child(4) > td > div > button')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_name('currentPassword').send_keys('qwer1234!')
    elm = driver.find_element_by_name('newPassword').send_keys('qwer1234!!')
    elm = driver.find_element_by_name('newPasswordConfirm').send_keys('qwer1234!!')
    sleep(2)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div/div[3]/button[1]')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._2Ki5-VQiUsW3itJyIauuvq > a > h1')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > button')
    btn_elm.click()
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
    btn_elm.click()
    sleep(2)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > button')
    btn_elm.click()
    sleep(2)
    # 변경한 비밀번호로 로그인
    elm = driver.find_element_by_id('username')
    elm.clear()
    elm.send_keys('yoojung_jang@tmax.co.kr')
    elm = driver.find_element_by_id('password').send_keys('qwer1234!!')
    btn_elm = driver.find_element_by_id('kc-form-buttons').click()
    sleep(2)
    try:
        elm = driver.find_element_by_id('username')
        ws['B7'] = '10'
        ws['C7'] = 'FAIL'
        return 0
        # 여기서 오류나면 끊어야됨 진행안됨
    except:
        ws['B7'] = '10'
        ws['C7'] = 'PASS'
        return 1

# 1
# Tmax Admin계정으로 HyperMeeting 계정으로 등록 확인
def test1():
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > a > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._19XtTBEROGh2PcaaoNIpB_ > div._2jgTBtopzy7N66IlhA1zM4 > div:nth-child(2) > button._2u7qp6n2F-I5dMK9oABuJG.VD_UJ7SvawulSczjIW4oU')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_id('searchId').send_keys('tmaxyoojung@gmail.com')
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[5]/button[1]')
    btn_elm.click()
    sleep(1)
    try:
        btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button')
        ws['B8'] = '1'
        ws['C8'] = 'FAIL'
        btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[1]/button')
        btn_elm.click()
        return 0
    except:
        ws['B8'] = '1'
        ws['C8'] = 'PASS'
        return 1

# 2
# Tmax Admin계정으로 신규 계정 생성하여 개별 등록
def test2(id):
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._19XtTBEROGh2PcaaoNIpB_ > div._2jgTBtopzy7N66IlhA1zM4 > div:nth-child(2) > button._2u7qp6n2F-I5dMK9oABuJG.VD_UJ7SvawulSczjIW4oU')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[5]/button[2]')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_name('username').send_keys('테스트')
    elm = driver.find_element_by_name('userId').send_keys(id)
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[2]/div/table/tr[2]/td/div[1]/button')
    btn_elm.click()
    elm = driver.find_element_by_name('password').send_keys('qwer1234!')
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
    btn_elm.click()
    sleep(1)
    try:
        btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[2]')
        btn_elm.click()
        ws['B9'] = '2'
        ws['C9'] = 'FAIL'
        sleep(1)
        btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div[3]/button[1]')
        btn_elm.click()
        return 0
    except:
        ws['B9'] = '2'
        ws['C9'] = 'PASS'
        btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div[3]/button[1]')
        btn_elm.click()
        return 1

# 435
# 4: Tmax Admin계정으로 auth에 가입된 계정 이름 일치하지 않게 등록 (일괄 등록)
# 3: Tmax Admin계정으로 auth에 가입된 계정 이름도 일치하게 등록 (일괄 등록)
# 5: Tmax Admin계정으로 신규 계정 생성하여 일괄 등록
def test435():
    # 4
    temp = 0
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._19XtTBEROGh2PcaaoNIpB_ > div._2jgTBtopzy7N66IlhA1zM4 > div:nth-child(2) > button:nth-child(1)')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div/div[2]/div[2]/div/label')
    btn_elm.click()
    sleep(1)
    # 이름 틀린 파일
    pyautogui.write("C:\\admin\\batch_register_tmax_admin.xlsx")
    pyautogui.hotkey('enter')
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div/div[2]/div[3]/button[1]')
    btn_elm.click()
    sleep(1)
    # 통합 계정 이름과 일치하지 않음이 떠야함
    try:
        elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div/div[3]/div[2]/div/table/tbody/tr/td[2][contains(text(), "이름")]')
        ws['B10'] = '4'
        ws['C10'] = 'PASS'
        temp = 1
    except:
        ws['B10'] = '4'
        ws['C10'] = 'FAIL'
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div/div[1]/button')
    btn_elm.click()
    sleep(1)

    # 3,5
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._19XtTBEROGh2PcaaoNIpB_ > div._2jgTBtopzy7N66IlhA1zM4 > div:nth-child(2) > button:nth-child(1)')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div/div[2]/div[2]/div/label')
    btn_elm.click()
    sleep(1)
    # 이름 같고 비번 틀린 계정과 새 계정 파일
    pyautogui.write("C:\\admin\\batch_register_tmax_admin2.xlsx")
    pyautogui.hotkey('enter')
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div/div[2]/div[3]/button[1]')
    btn_elm.click()
    sleep(1)
    # 성공 두개
    try:
        elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div/div[2]/div/table/tr[3]/td[contains(text(), "0")]')
        ws['B11'] = '3,5'
        ws['C11'] = 'PASS'
        temp = temp + 1
    except:
        ws['B11'] = '3,5'
        ws['C11'] = 'FAIL'
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div/div[1]/button')
    btn_elm.click()
    sleep(1)
    return temp

# 7
# Tmax Admin계정으로 B2B 구매 계정 승인
def test7():
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._1OwJclDo5JkkZuvbXf3bOZ > div._3BshymDCc0OSBa0oS0zJj- > div > div > div')
    btn_elm.click()
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._1OwJclDo5JkkZuvbXf3bOZ > div._3BshymDCc0OSBa0oS0zJj- > div > div > div > div._2taOCr9V0CnonoT0H_eksA > span:nth-child(2)')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_id('searchName').send_keys('tmaxyoojung@gmail.com')
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._1OwJclDo5JkkZuvbXf3bOZ > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._19XtTBEROGh2PcaaoNIpB_ > div._1ICwZfGa-u0vwVoGlQwVkD > table > tbody > tr > td:nth-child(4) > button')
    btn_elm.click()
    sleep(1)
    # 전화번호
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div:nth-child(4) > table > tr:nth-child(4) > td > div > div > div._1eqdx1TRXvnj_yZ3A2CnJ3._2VoB-BqDjcx41XiO-Sq6iJ')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div:nth-child(4) > table > tr:nth-child(4) > td > div > div > div._2taOCr9V0CnonoT0H_eksA > span:nth-child(1)')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_name('phoneNumberMiddle').send_keys('1234')
    elm = driver.find_element_by_name('phoneNumberEnd').send_keys('1234')
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div:nth-child(4) > table > tr:nth-child(5) > td > div > div > div._1eqdx1TRXvnj_yZ3A2CnJ3._2VoB-BqDjcx41XiO-Sq6iJ')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div:nth-child(4) > table > tr:nth-child(5) > td > div > div > div._2taOCr9V0CnonoT0H_eksA > span:nth-child(1)')
    btn_elm.click()
    sleep(1)
    # 시작일
    elm = driver.find_element_by_name('validStartDate').send_keys('2021')
    pyautogui.hotkey('right')
    pyautogui.write('02')
    pyautogui.hotkey('right')
    pyautogui.write('10')
    sleep(1)
    # 종료일
    elm = driver.find_element_by_name('validEndDate').send_keys('2021')
    pyautogui.hotkey('right')
    pyautogui.write('12')
    pyautogui.hotkey('right')
    pyautogui.write('20')
    sleep(1)
    elm = driver.find_element_by_name('hostCount').send_keys('10')
    elm = driver.find_element_by_name('companyName').send_keys('ccc')
    elm = driver.find_element_by_name('salesRepresentative').send_keys('나')
    sleep(1)
    try:
        btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div._1kbII_WxXJa1Ewh8pE8MFE > button._2u7qp6n2F-I5dMK9oABuJG')
        btn_elm.click()
        btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
        btn_elm.click()
        ws['B12'] = '7'
        ws['C12'] = 'PASS'
        return 1
    except:
        ws['B12'] = '7'
        ws['C12'] = 'FAIL'
        btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div._1kbII_WxXJa1Ewh8pE8MFE > button._12i0zizoctRi7HO79lAz18')
        btn_elm.click()
        btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
        btn_elm.click()
        return 0

# 8
# Tmax Admin계정으로 B2B 구매만료 계정 승인
def test8(id):
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._1OwJclDo5JkkZuvbXf3bOZ > div._3BshymDCc0OSBa0oS0zJj- > div > div > div')
    btn_elm.click()
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._1OwJclDo5JkkZuvbXf3bOZ > div._3BshymDCc0OSBa0oS0zJj- > div > div > div > div._2taOCr9V0CnonoT0H_eksA > span:nth-child(2)')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_id('searchName').send_keys(id)
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._1OwJclDo5JkkZuvbXf3bOZ > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._19XtTBEROGh2PcaaoNIpB_ > div._1ICwZfGa-u0vwVoGlQwVkD > table > tbody > tr > td:nth-child(4) > button')
    btn_elm.click()
    sleep(1)
    # 전화번호
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div:nth-child(4) > table > tr:nth-child(4) > td > div > div > div._1eqdx1TRXvnj_yZ3A2CnJ3._2VoB-BqDjcx41XiO-Sq6iJ')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div:nth-child(4) > table > tr:nth-child(4) > td > div > div > div._2taOCr9V0CnonoT0H_eksA > span:nth-child(1)')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_name('phoneNumberMiddle').send_keys('1234')
    elm = driver.find_element_by_name('phoneNumberEnd').send_keys('1234')
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div:nth-child(4) > table > tr:nth-child(5) > td > div > div > div._1eqdx1TRXvnj_yZ3A2CnJ3._2VoB-BqDjcx41XiO-Sq6iJ')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div:nth-child(4) > table > tr:nth-child(5) > td > div > div > div._2taOCr9V0CnonoT0H_eksA > span:nth-child(1)')
    btn_elm.click()
    sleep(1)
    # 시작일
    elm = driver.find_element_by_name('validStartDate').send_keys('2021')
    pyautogui.hotkey('right')
    pyautogui.write('02')
    pyautogui.hotkey('right')
    pyautogui.write('10')
    sleep(1)
    # 종료일
    elm = driver.find_element_by_name('validEndDate').send_keys('2021')
    pyautogui.hotkey('right')
    pyautogui.write('02')
    pyautogui.hotkey('right')
    pyautogui.write('20')
    sleep(1)
    elm = driver.find_element_by_name('hostCount').send_keys('10')
    elm = driver.find_element_by_name('companyName').send_keys('cccc')
    elm = driver.find_element_by_name('salesRepresentative').send_keys('나')
    sleep(1)
    try:
        btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div._1kbII_WxXJa1Ewh8pE8MFE > button._2u7qp6n2F-I5dMK9oABuJG')
        btn_elm.click()
        btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
        btn_elm.click()
        ws['B13'] = '8'
        ws['C13'] = 'PASS'
        return 1
    except:
        ws['B13'] = '8'
        ws['C13'] = 'FAIL'
        btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div._1kbII_WxXJa1Ewh8pE8MFE > button._12i0zizoctRi7HO79lAz18')
        btn_elm.click()
        btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
        btn_elm.click()
        return 0

# 9
# Tmax Admin계정으로 B2C 구매 계정 승인
def test9(id):
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._1OwJclDo5JkkZuvbXf3bOZ > div._3BshymDCc0OSBa0oS0zJj- > div > div > div')
    btn_elm.click()
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._1OwJclDo5JkkZuvbXf3bOZ > div._3BshymDCc0OSBa0oS0zJj- > div > div > div > div._2taOCr9V0CnonoT0H_eksA > span:nth-child(2)')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_id('searchName').send_keys(id)
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._1OwJclDo5JkkZuvbXf3bOZ > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._19XtTBEROGh2PcaaoNIpB_ > div._1ICwZfGa-u0vwVoGlQwVkD > table > tbody > tr > td:nth-child(4) > button')
    btn_elm.click()
    sleep(1)
    # B2C 계정
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div:nth-child(2) > table > tr > td > label:nth-child(2)')
    btn_elm.click()
    # 전화번호
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div:nth-child(4) > table > tr:nth-child(4) > td > div > div > div._1eqdx1TRXvnj_yZ3A2CnJ3._2VoB-BqDjcx41XiO-Sq6iJ')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div:nth-child(4) > table > tr:nth-child(4) > td > div > div > div._2taOCr9V0CnonoT0H_eksA > span:nth-child(1)')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_name('phoneNumberMiddle').send_keys('1234')
    elm = driver.find_element_by_name('phoneNumberEnd').send_keys('1234')
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div:nth-child(4) > table > tr:nth-child(5) > td > div > div > div._1eqdx1TRXvnj_yZ3A2CnJ3._2VoB-BqDjcx41XiO-Sq6iJ')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div:nth-child(4) > table > tr:nth-child(5) > td > div > div > div._2taOCr9V0CnonoT0H_eksA > span:nth-child(1)')
    btn_elm.click()
    sleep(1)
    try:
        btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div._1kbII_WxXJa1Ewh8pE8MFE > button._2u7qp6n2F-I5dMK9oABuJG')
        btn_elm.click()
        btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
        btn_elm.click()
        ws['B14'] = '9'
        ws['C14'] = 'PASS'
        return 1
    except:
        ws['B14'] = '9'
        ws['C14'] = 'FAIL'
        btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div._1kbII_WxXJa1Ewh8pE8MFE > button._12i0zizoctRi7HO79lAz18')
        btn_elm.click()
        btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
        btn_elm.click()
        return 0

# 614
# 6: TC3에서 패스워드 다르게 등록한 계정 로그인 확인
# 14: 비구매 계정으로 회의실 만들기 팝업창이 뜨는지 확인
def test614():
    # 6
    temp = 0
    btn_elm = driver.find_element_by_css_selector('#root > div > div._2Ki5-VQiUsW3itJyIauuvq > a > h1')
    btn_elm.click()
    sleep(1)
    # 로그아웃
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
    btn_elm.click()
    sleep(1)
    # 로그인
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > button')
    btn_elm.click()
    sleep(2)
    elm = driver.find_element_by_id('username')
    elm.clear()
    elm.send_keys('admintest0@tmax.co.kr')
    elm = driver.find_element_by_id('password').send_keys('qwer1234!')
    btn_elm = driver.find_element_by_id('kc-form-buttons').click()
    sleep(2)
    try:
        elm = driver.find_element_by_id('username')
        ws['B15'] = '6'
        ws['C15'] = 'FAIL'
        # 여기서 오류나면 끊어야됨 진행안됨
    except:
        ws['B15'] = '6'
        ws['C15'] = 'PASS'
        temp = 1

    # 14
    # 동의 창 뜨는거
    btn_elm = driver.find_element_by_css_selector('#root > div > div._22JcKHWNNh84WFExow3bG_ > div > div._1-jcAvHaWoHo5zNTWt_fTs > p._1kG9gUZ7Lun75q6APCcXWc.dxtBZiRsSyLi2iqyrMZla > label')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._22JcKHWNNh84WFExow3bG_ > div > div.xlYyTVHR4ofNupcHZ3A3F > button._1565LCtRGZmVxIV5arpcDX.false')
    btn_elm.click()
    sleep(1)

    btn_elm = driver.find_element_by_css_selector('#root > div > div._227SmuIIA_SGYBpxZc55bj > div._11HDw8Md_fZm5HmzzPFclC > button')
    btn_elm.click()
    sleep(1)
    try:
        elm = driver.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[2]/div/p[1]/span[contains(text(), "구매 문의")]')
        ws['B16'] = '14'
        ws['C16'] = 'PASS'
        btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[1]/button')
        btn_elm.click()
        temp = temp + 1
        return temp
    except:
        ws['B16'] = '14'
        ws['C16'] = 'FAIL'
        btn_elm = driver.find_element_by_xpath('/html/body/div/div/div/div[1]/button')
        btn_elm.click()
        return temp

# 24
# 고객 Admin 구매 계정으로 회의실 만들기 팝업창이 뜨지 않는지 확인
def test24():
    # 로그아웃
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
    btn_elm.click()
    sleep(1)
    # 로그인
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > button')
    btn_elm.click()
    sleep(2)
    elm = driver.find_element_by_id('username')
    elm.clear()
    elm.send_keys('tmaxyoojung@gmail.com')
    elm = driver.find_element_by_id('password').send_keys('qwer1234!')
    btn_elm = driver.find_element_by_id('kc-form-buttons').click()
    sleep(2)
    # 동의 창 뜨는거
    btn_elm = driver.find_element_by_css_selector('#root > div > div._22JcKHWNNh84WFExow3bG_ > div > div._1-jcAvHaWoHo5zNTWt_fTs > p._1kG9gUZ7Lun75q6APCcXWc.dxtBZiRsSyLi2iqyrMZla > label')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._22JcKHWNNh84WFExow3bG_ > div > div.xlYyTVHR4ofNupcHZ3A3F > button._1565LCtRGZmVxIV5arpcDX.false')
    btn_elm.click()
    sleep(1)

    btn_elm = driver.find_element_by_css_selector('#root > div > div._227SmuIIA_SGYBpxZc55bj > div._11HDw8Md_fZm5HmzzPFclC > button')
    btn_elm.click()
    sleep(2)
    try:
        elm = driver.find_element_by_id('input_room_name')
        ws['B17'] = '24'
        ws['C17'] = 'PASS'
        btn_elm = driver.find_element_by_xpath('/html/body/div/div/div/div[1]/button')
        btn_elm.click()
        return 1
    except:
        ws['B17'] = '24'
        ws['C17'] = 'FAIL'
        btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[1]/button')
        btn_elm.click()
        return 0

# 16
# 고객 Admin계정으로 auth에 가입된 계정 이름 일치하지 않게 등록 (호스트 개별 등록)
def test16():
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > a > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div.ICjgBEGgF_4nnT6DOdQ9t > div > dl:nth-child(2) > dd:nth-child(2) > a > button')
    btn_elm.click()
    sleep(1)
    # 개별 등록
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._19XtTBEROGh2PcaaoNIpB_ > div._2jgTBtopzy7N66IlhA1zM4 > div:nth-child(2) > button._2u7qp6n2F-I5dMK9oABuJG.VD_UJ7SvawulSczjIW4oU')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_name('username').send_keys('테스트')
    elm = driver.find_element_by_name('userId').send_keys('admintest00@tmax.co.kr')
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div.RH4-LqVZJjx9RYchbxlwH > table > tr:nth-child(3) > td > button')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_name('password').send_keys('qwer1234@')
    elm = driver.find_element_by_name('department').send_keys('부서')
    elm = driver.find_element_by_name('jobTitle').send_keys('직함')
    sleep(1)
    # 전화번호
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div.RH4-LqVZJjx9RYchbxlwH > table > tr:nth-child(7) > td > div > div > div._1eqdx1TRXvnj_yZ3A2CnJ3._2VoB-BqDjcx41XiO-Sq6iJ')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div.RH4-LqVZJjx9RYchbxlwH > table > tr:nth-child(7) > td > div > div > div._2taOCr9V0CnonoT0H_eksA > span:nth-child(1)')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_name('phoneNumberMiddle').send_keys('1234')
    elm = driver.find_element_by_name('phoneNumberEnd').send_keys('1234')
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div._1kbII_WxXJa1Ewh8pE8MFE > button._2u7qp6n2F-I5dMK9oABuJG')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
    btn_elm.click()
    sleep(1)
    try:
        elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[1]/p[contains(text(), "등록 실패")]')
        ws['B18'] = '16'
        ws['C18'] = 'PASS'
        btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button')
        btn_elm.click()
        return 1
    except:
        ws['B18'] = '16'
        ws['C18'] = 'FAIL'
        # 원래 뭐 없었는데 일단 지금 현상태로 넣음 이슈때문에
        btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div._1kbII_WxXJa1Ewh8pE8MFE > button._12i0zizoctRi7HO79lAz18')
        btn_elm.click()
        sleep(1)
        btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
        btn_elm.click()
        sleep(1)
        return 0

# 15
# 고객 Admin계정으로 auth에 가입된 계정 이름도 일치하게 등록 (호스트 개별 등록)
def test15():
    # 개별 등록
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._19XtTBEROGh2PcaaoNIpB_ > div._2jgTBtopzy7N66IlhA1zM4 > div:nth-child(2) > button._2u7qp6n2F-I5dMK9oABuJG.VD_UJ7SvawulSczjIW4oU')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_name('username').send_keys('테스트00')
    elm = driver.find_element_by_name('userId').send_keys('admintest00@tmax.co.kr')
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div.RH4-LqVZJjx9RYchbxlwH > table > tr:nth-child(3) > td > button')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_name('password').send_keys('qwer1234@')
    elm = driver.find_element_by_name('department').send_keys('부서')
    elm = driver.find_element_by_name('jobTitle').send_keys('직함')
    sleep(1)
    # 전화번호
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div.RH4-LqVZJjx9RYchbxlwH > table > tr:nth-child(7) > td > div > div > div._1eqdx1TRXvnj_yZ3A2CnJ3._2VoB-BqDjcx41XiO-Sq6iJ')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div.RH4-LqVZJjx9RYchbxlwH > table > tr:nth-child(7) > td > div > div > div._2taOCr9V0CnonoT0H_eksA > span:nth-child(1)')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_name('phoneNumberMiddle').send_keys('1234')
    elm = driver.find_element_by_name('phoneNumberEnd').send_keys('1234')
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div._1kbII_WxXJa1Ewh8pE8MFE > button._2u7qp6n2F-I5dMK9oABuJG')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
    btn_elm.click()
    sleep(1)
    try:
        elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[1]/p[contains(text(), "통합")]')
        ws['B19'] = '15'
        ws['C19'] = 'PASS'
        btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button')
        btn_elm.click()
        return 1
    except:
        ws['B19'] = '15'
        ws['C19'] = 'FAIL'
        return 0

# 17
# 고객 Admin계정으로 신규 계정 생성하여 호스트 개별 등록
def test17(id):
    # 개별 등록
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._19XtTBEROGh2PcaaoNIpB_ > div._2jgTBtopzy7N66IlhA1zM4 > div:nth-child(2) > button._2u7qp6n2F-I5dMK9oABuJG.VD_UJ7SvawulSczjIW4oU')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_name('username').send_keys('테스트')
    elm = driver.find_element_by_name('userId').send_keys(id)
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div.RH4-LqVZJjx9RYchbxlwH > table > tr:nth-child(3) > td > button')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_name('password').send_keys('qwer1234!')
    elm = driver.find_element_by_name('department').send_keys('부서')
    elm = driver.find_element_by_name('jobTitle').send_keys('직함')
    sleep(1)
    # 전화번호
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div.RH4-LqVZJjx9RYchbxlwH > table > tr:nth-child(7) > td > div > div > div._1eqdx1TRXvnj_yZ3A2CnJ3._2VoB-BqDjcx41XiO-Sq6iJ')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div.RH4-LqVZJjx9RYchbxlwH > table > tr:nth-child(7) > td > div > div > div._2taOCr9V0CnonoT0H_eksA > span:nth-child(1)')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_name('phoneNumberMiddle').send_keys('1234')
    elm = driver.find_element_by_name('phoneNumberEnd').send_keys('1234')
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div._1kbII_WxXJa1Ewh8pE8MFE > button._2u7qp6n2F-I5dMK9oABuJG')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
    btn_elm.click()
    sleep(1)
    try:
        elm = driver.find_element_by_name('username')
        ws['B20'] = '17'
        ws['C20'] = 'FAIL'
        btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div._1kbII_WxXJa1Ewh8pE8MFE > button._12i0zizoctRi7HO79lAz18')
        btn_elm.click()
        sleep(1)
        btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
        btn_elm.click()
        sleep(1)
        return 0
    except:
        ws['B20'] = '17'
        ws['C20'] = 'PASS'
        return 1

# 201921
# 20: 고객 Admin계정으로 auth에 가입된 계정 이름 일치하지 않게 등록 (호스트 일괄 등록)
# 19: 고객 Admin계정으로 auth에 가입된 계정 이름도 일치하게 등록 (호스트 일괄 등록)
# 21: 고객 Admin계정으로 신규 계정 생성하여 호스트 일괄 등록
def test201921():
    # 20
    temp = 0
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._19XtTBEROGh2PcaaoNIpB_ > div._2jgTBtopzy7N66IlhA1zM4 > div:nth-child(2) > button:nth-child(2)')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div/div[2]/div[2]/div/label')
    btn_elm.click()
    sleep(1)
    # 이름 틀린 파일
    pyautogui.write("C:\\admin\\batch_register_customer_admin.xlsx")
    pyautogui.hotkey('enter')
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div/div[2]/div[3]/button[1]')
    btn_elm.click()
    sleep(1)
    # 통합 계정 이름과 일치하지 않음이 떠야함
    try:
        elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div/div[3]/div[2]/div/table/tbody/tr/td[2][contains(text(), "이름")]')
        ws['B21'] = '20'
        ws['C21'] = 'PASS'
        temp = 1
    except:
        ws['B21'] = '20'
        ws['C21'] = 'FAIL'
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div/div[1]/button')
    btn_elm.click()
    sleep(1)

    # 19,21
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._19XtTBEROGh2PcaaoNIpB_ > div._2jgTBtopzy7N66IlhA1zM4 > div:nth-child(2) > button:nth-child(2)')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div/div[2]/div[2]/div/label')
    btn_elm.click()
    sleep(1)
    # 이름 같고 비번 틀린 계정과 새 계정 파일
    pyautogui.write("C:\\admin\\batch_register_customer_admin2.xlsx")
    pyautogui.hotkey('enter')
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div/div[2]/div[3]/button[1]')
    btn_elm.click()
    sleep(1)
    # 성공 두개
    try:
        elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div/div[2]/div/table/tr[3]/td[contains(text(), "0")]')
        ws['B22'] = '19,21'
        ws['C22'] = 'PASS'
        temp = temp + 1
    except:
        ws['B22'] = '19,21'
        ws['C22'] = 'FAIL'
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div/div[1]/button')
    btn_elm.click()
    sleep(1)
    return temp

# 25
# 고객 Admin 구매만료 계정으로 회의실 만들기 팝업창이 뜨는지 확인
def test25(id):
    btn_elm = driver.find_element_by_css_selector('#root > div > div._2Ki5-VQiUsW3itJyIauuvq > a > h1')
    btn_elm.click()
    sleep(1)
    # 로그아웃
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
    btn_elm.click()
    sleep(1)
    # 로그인
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > button')
    btn_elm.click()
    sleep(2)
    elm = driver.find_element_by_id('username')
    elm.clear()
    elm.send_keys(id)
    elm = driver.find_element_by_id('password').send_keys('qwer1234!')
    btn_elm = driver.find_element_by_id('kc-form-buttons').click()
    sleep(2)
    # 동의 창 뜨는거
    btn_elm = driver.find_element_by_css_selector('#root > div > div._22JcKHWNNh84WFExow3bG_ > div > div._1-jcAvHaWoHo5zNTWt_fTs > p._1kG9gUZ7Lun75q6APCcXWc.dxtBZiRsSyLi2iqyrMZla > label')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._22JcKHWNNh84WFExow3bG_ > div > div.xlYyTVHR4ofNupcHZ3A3F > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._22JcKHWNNh84WFExow3bG_ > div > div._1-jcAvHaWoHo5zNTWt_fTs > p._1kG9gUZ7Lun75q6APCcXWc.dxtBZiRsSyLi2iqyrMZla > label')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._22JcKHWNNh84WFExow3bG_ > div > div.xlYyTVHR4ofNupcHZ3A3F > button._1565LCtRGZmVxIV5arpcDX.false')
    btn_elm.click()
    sleep(1)

    btn_elm = driver.find_element_by_css_selector('#root > div > div._227SmuIIA_SGYBpxZc55bj > div._11HDw8Md_fZm5HmzzPFclC > button')
    btn_elm.click()
    sleep(1)
    try:
        elm = driver.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[2]/div/p[1]/span[contains(text(), "구매 문의")]')
        ws['B23'] = '25'
        ws['C23'] = 'PASS'
        btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[1]/button')
        btn_elm.click()
        return 1
    except:
        ws['B23'] = '25'
        ws['C23'] = 'FAIL'
        btn_elm = driver.find_element_by_xpath('/html/body/div/div/div/div[1]/button')
        btn_elm.click()
        return 0

# 27
# 고객 Admin 구매만료 계정에 등록된 호스트 계정으로 회의실 만들기 팝업창이 뜨는지 확인
def test27():
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > a > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div.ICjgBEGgF_4nnT6DOdQ9t > div > dl:nth-child(2) > dd:nth-child(2) > a > button')
    btn_elm.click()
    sleep(1)
    # 개별 등록
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._19XtTBEROGh2PcaaoNIpB_ > div._2jgTBtopzy7N66IlhA1zM4 > div:nth-child(2) > button._2u7qp6n2F-I5dMK9oABuJG.VD_UJ7SvawulSczjIW4oU')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_name('username').send_keys('테스트0')
    elm = driver.find_element_by_name('userId').send_keys('admintest0@tmax.co.kr')
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div.RH4-LqVZJjx9RYchbxlwH > table > tr:nth-child(3) > td > button')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_name('password').send_keys('qwer1234!')
    elm = driver.find_element_by_name('department').send_keys('부서')
    elm = driver.find_element_by_name('jobTitle').send_keys('직함')
    sleep(1)
    # 전화번호
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div.RH4-LqVZJjx9RYchbxlwH > table > tr:nth-child(7) > td > div > div > div._1eqdx1TRXvnj_yZ3A2CnJ3._2VoB-BqDjcx41XiO-Sq6iJ')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div.RH4-LqVZJjx9RYchbxlwH > table > tr:nth-child(7) > td > div > div > div._2taOCr9V0CnonoT0H_eksA > span:nth-child(1)')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_name('phoneNumberMiddle').send_keys('1234')
    elm = driver.find_element_by_name('phoneNumberEnd').send_keys('1234')
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div._1kbII_WxXJa1Ewh8pE8MFE > button._2u7qp6n2F-I5dMK9oABuJG')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._2Ki5-VQiUsW3itJyIauuvq > a > h1')
    btn_elm.click()
    sleep(1)
    # 로그아웃
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
    btn_elm.click()
    sleep(1)
    # 로그인
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > button')
    btn_elm.click()
    sleep(2)
    elm = driver.find_element_by_id('username')
    elm.clear()
    elm.send_keys('admintest0@tmax.co.kr')
    elm = driver.find_element_by_id('password').send_keys('qwer1234!')
    btn_elm = driver.find_element_by_id('kc-form-buttons').click()
    sleep(2)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._227SmuIIA_SGYBpxZc55bj > div._11HDw8Md_fZm5HmzzPFclC > button')
    btn_elm.click()
    sleep(1)
    try:
        elm = driver.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[2]/div/p[1]/span[contains(text(), "구매 문의")]')
        ws['B24'] = '27'
        ws['C24'] = 'PASS'
        btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[1]/button')
        btn_elm.click()
        return 1
    except:
        ws['B24'] = '27'
        ws['C24'] = 'FAIL'
        btn_elm = driver.find_element_by_xpath('/html/body/div/div/div/div[1]/button')
        btn_elm.click()
        return 0

# 1826
# 18: TC14에서 패스워드 다르게 등록한 계정 로그인 확인
# 26: 고객 Admin 구매 계정에 등록된 호스트 계정으로 회의실 만들기 팝업창이 뜨지 않는지 확인
def test1826():
    # 18
    temp = 0
    # 로그아웃
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
    btn_elm.click()
    sleep(1)
    # 로그인
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > button')
    btn_elm.click()
    sleep(2)
    elm = driver.find_element_by_id('username')
    elm.clear()
    elm.send_keys('admintest00@tmax.co.kr')
    elm = driver.find_element_by_id('password').send_keys('qwer1234!')
    btn_elm = driver.find_element_by_id('kc-form-buttons').click()
    sleep(2)
    try:
        elm = driver.find_element_by_id('username')
        ws['B25'] = '18'
        ws['C25'] = 'FAIL'
        # 여기서 오류나면 끊어야됨 진행안됨
    except:
        ws['B25'] = '18'
        ws['C25'] = 'PASS'
        temp = 1

    # 26
    # 동의 창 뜨는거
    btn_elm = driver.find_element_by_css_selector('#root > div > div._22JcKHWNNh84WFExow3bG_ > div > div._1-jcAvHaWoHo5zNTWt_fTs > p._1kG9gUZ7Lun75q6APCcXWc.dxtBZiRsSyLi2iqyrMZla > label')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._22JcKHWNNh84WFExow3bG_ > div > div.xlYyTVHR4ofNupcHZ3A3F > button._1565LCtRGZmVxIV5arpcDX.false')
    btn_elm.click()
    sleep(1)

    btn_elm = driver.find_element_by_css_selector('#root > div > div._227SmuIIA_SGYBpxZc55bj > div._11HDw8Md_fZm5HmzzPFclC > button')
    btn_elm.click()
    sleep(1)
    try:
        elm = driver.find_element_by_id('input_room_name')
        ws['B26'] = '26'
        ws['C26'] = 'PASS'
        btn_elm = driver.find_element_by_xpath('/html/body/div/div/div/div[1]/button')
        btn_elm.click()
        temp = temp + 1
        return temp
    except:
        ws['B26'] = '26'
        ws['C26'] = 'FAIL'
        btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[1]/button')
        btn_elm.click()
        return temp

# 22
# TC19에서 패스워드 다르게 등록한 계정 로그인 확인
def test22():
    # 로그아웃
    temp = 0
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
    btn_elm.click()
    sleep(1)
    # 로그인
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > button')
    btn_elm.click()
    sleep(2)
    elm = driver.find_element_by_id('username')
    elm.clear()
    elm.send_keys('admintest000@tmax.co.kr')
    elm = driver.find_element_by_id('password').send_keys('qwer1234!')
    btn_elm = driver.find_element_by_id('kc-form-buttons').click()
    sleep(2)
    try:
        elm = driver.find_element_by_id('username')
        ws['B27'] = '22'
        ws['C27'] = 'FAIL'
        # 여기서 오류나면 끊어야됨 진행안됨
    except:
        ws['B27'] = '22'
        ws['C27'] = 'PASS'
        temp = 1
    # 동의 창 뜨는거 - 얘만 제대로 작동
    btn_elm = driver.find_element_by_css_selector('#root > div > div._22JcKHWNNh84WFExow3bG_ > div > div._1-jcAvHaWoHo5zNTWt_fTs > p._1kG9gUZ7Lun75q6APCcXWc.dxtBZiRsSyLi2iqyrMZla > label')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._22JcKHWNNh84WFExow3bG_ > div > div.xlYyTVHR4ofNupcHZ3A3F > button._1565LCtRGZmVxIV5arpcDX.false')
    btn_elm.click()
    sleep(1)
    return temp

# 28
# B2C 구매 계정으로 회의실 만들기 팝업창이 뜨지 않는지 확인
def test28(id):
    # 로그아웃
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
    btn_elm.click()
    sleep(1)
    # 로그인
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > button')
    btn_elm.click()
    sleep(2)
    elm = driver.find_element_by_id('username')
    elm.clear()
    elm.send_keys(id)
    elm = driver.find_element_by_id('password').send_keys('qwer1234!')
    btn_elm = driver.find_element_by_id('kc-form-buttons').click()
    sleep(2)
    # 동의 창 뜨는거
    btn_elm = driver.find_element_by_css_selector('#root > div > div._22JcKHWNNh84WFExow3bG_ > div > div._1-jcAvHaWoHo5zNTWt_fTs > p._1kG9gUZ7Lun75q6APCcXWc.dxtBZiRsSyLi2iqyrMZla > label')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._22JcKHWNNh84WFExow3bG_ > div > div.xlYyTVHR4ofNupcHZ3A3F > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._22JcKHWNNh84WFExow3bG_ > div > div._1-jcAvHaWoHo5zNTWt_fTs > p._1kG9gUZ7Lun75q6APCcXWc.dxtBZiRsSyLi2iqyrMZla > label')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._22JcKHWNNh84WFExow3bG_ > div > div.xlYyTVHR4ofNupcHZ3A3F > button._1565LCtRGZmVxIV5arpcDX.false')
    btn_elm.click()
    sleep(1)

    btn_elm = driver.find_element_by_css_selector('#root > div > div._227SmuIIA_SGYBpxZc55bj > div._11HDw8Md_fZm5HmzzPFclC > button')
    btn_elm.click()
    sleep(1)
    try:
        elm = driver.find_element_by_id('input_room_name')
        ws['B28'] = '28'
        ws['C28'] = 'PASS'
        btn_elm = driver.find_element_by_xpath('/html/body/div/div/div/div[1]/button')
        btn_elm.click()
        return 1
    except:
        ws['B28'] = '28'
        ws['C28'] = 'FAIL'
        btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[1]/button')
        btn_elm.click()
        return 0

# 23
# 고객 Admin 구매 계정으로 등록한 호스트가 삭제되는지 확인
def test23():
    # 로그아웃
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
    btn_elm.click()
    sleep(1)
    # 로그인
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > button')
    btn_elm.click()
    sleep(2)
    elm = driver.find_element_by_id('username')
    elm.clear()
    elm.send_keys('tmaxyoojung@gmail.com')
    elm = driver.find_element_by_id('password').send_keys('qwer1234!')
    btn_elm = driver.find_element_by_id('kc-form-buttons').click()
    sleep(2)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > a > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div.ICjgBEGgF_4nnT6DOdQ9t > div > dl:nth-child(2) > dd:nth-child(2) > a > button')
    btn_elm.click()
    sleep(1)
    # 이름으로 검색
    elm = driver.find_element_by_id('searchName').send_keys('테스트000')
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._1OwJclDo5JkkZuvbXf3bOZ > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._19XtTBEROGh2PcaaoNIpB_ > div._1ICwZfGa-u0vwVoGlQwVkD > table > tbody > tr > td:nth-child(2)')
    btn_elm.click()
    sleep(1)
    # 삭제
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div._1ScRXCyfLzSgOrfscbk2CO > div._1kbII_WxXJa1Ewh8pE8MFE > button:nth-child(3)')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
    btn_elm.click()
    sleep(1)
    # 삭제 확인
    elm = driver.find_element_by_id('searchName')
    elm.clear()
    elm.send_keys('테스트000')
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._1OwJclDo5JkkZuvbXf3bOZ > button')
    btn_elm.click()
    sleep(1)
    try:
        elm = driver.find_element_by_xpath('/html/body/div[1]/div/div[2]/div[2]/div/div[2]/div/div/div[2]/div[2]/table/tbody/tr/td[contains(text(), "없습니다.")]')
        ws['B29'] = '23'
        ws['C29'] = 'PASS'
        return 1
    except:
        ws['B29'] = '23'
        ws['C29'] = 'FAIL'
        return 0

# 12
# Tmax Admin 계정으로 승인한 B2B 구매 계정과 그 계정의 호스트들이 삭제되는지 확인
def test12():
    temp = 0
    btn_elm = driver.find_element_by_css_selector('#root > div > div._2Ki5-VQiUsW3itJyIauuvq > a > h1')
    btn_elm.click()
    sleep(1)
    # 로그아웃
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
    btn_elm.click()
    sleep(1)
    # 로그인
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > button')
    btn_elm.click()
    sleep(2)
    elm = driver.find_element_by_id('username')
    elm.clear()
    elm.send_keys('yoojung_jang@tmax.co.kr')
    # 10에서 비번 바꾼걸로
    elm = driver.find_element_by_id('password').send_keys('qwer1234!!')
    btn_elm = driver.find_element_by_id('kc-form-buttons').click()
    sleep(2)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._1MJOC3uRPv7hN-Ttr3e9ap > a > button')
    btn_elm.click()
    sleep(1)
    # 구매 고객
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div.ICjgBEGgF_4nnT6DOdQ9t > div > dl:nth-child(1) > dd:nth-child(3) > a > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(3) > div._1OwJclDo5JkkZuvbXf3bOZ > div._3BshymDCc0OSBa0oS0zJj- > div > div > div > div._1eqdx1TRXvnj_yZ3A2CnJ3._2VoB-BqDjcx41XiO-Sq6iJ')
    btn_elm.click()
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(3) > div._1OwJclDo5JkkZuvbXf3bOZ > div._3BshymDCc0OSBa0oS0zJj- > div > div > div > div._2taOCr9V0CnonoT0H_eksA > span:nth-child(3)')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_id('searchName').send_keys('tmaxyoojung@gmail.com')
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(3) > div._1OwJclDo5JkkZuvbXf3bOZ > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(3) > div._19XtTBEROGh2PcaaoNIpB_ > div._1ICwZfGa-u0vwVoGlQwVkD > table > tbody > tr > td:nth-child(3)')
    btn_elm.click()
    sleep(1)
    # 삭제
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div._1ScRXCyfLzSgOrfscbk2CO > div._1kbII_WxXJa1Ewh8pE8MFE > button:nth-child(3)')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
    btn_elm.click()
    sleep(1)
    # 고객 등록 및 조회
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div.ICjgBEGgF_4nnT6DOdQ9t > div > dl:nth-child(1) > dd:nth-child(2) > a > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._1OwJclDo5JkkZuvbXf3bOZ > div._3BshymDCc0OSBa0oS0zJj- > div > div > div')
    btn_elm.click()
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._1OwJclDo5JkkZuvbXf3bOZ > div._3BshymDCc0OSBa0oS0zJj- > div > div > div > div._2taOCr9V0CnonoT0H_eksA > span:nth-child(2)')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_id('searchName').send_keys('tmaxyoojung@gmail.com')
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._1OwJclDo5JkkZuvbXf3bOZ > button')
    btn_elm.click()
    sleep(1)
    try:
        elm = driver.find_element_by_xpath('/html/body/div/div/div[2]/div[2]/div/div[2]/div/div/div[2]/div[2]/table/tbody/tr/td[contains(text(), "없습니다.")]')
        ws['B30'] = '12'
        ws['C30'] = 'FAIL'
    except:
        ws['B30'] = '12'
        ws['C30'] = 'PASS'
        temp = 1
    # 호스트도 함께 삭제됐나 확인
    elm = driver.find_element_by_id('searchName')
    elm.clear()
    elm.send_keys('admintest00@tmax.co.kr')
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._1OwJclDo5JkkZuvbXf3bOZ > button')
    btn_elm.click()
    sleep(1)
    try:
        elm = driver.find_element_by_xpath('/html/body/div/div/div[2]/div[2]/div/div[2]/div/div/div[2]/div[2]/table/tbody/tr/td[contains(text(), "없습니다.")]')
        ws['B31'] = '12(host)'
        ws['C31'] = 'FAIL'
        return temp
    except:
        ws['B31'] = '12(host)'
        ws['C31'] = 'PASS'
        temp = temp + 1
        return temp

# 13
# Tmax Admin 계정으로 승인한 B2C 구매 계정이 삭제되는지 확인
def test13(id):
    # 구매 고객
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div.ICjgBEGgF_4nnT6DOdQ9t > div > dl:nth-child(1) > dd:nth-child(3) > a > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div.Sfeh2b_LZlMgGRxpm7woT > button._18ke5wDxl9pYTcs70Gnwr3.false')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(3) > div._1OwJclDo5JkkZuvbXf3bOZ > div._3BshymDCc0OSBa0oS0zJj- > div > div > div > div._1eqdx1TRXvnj_yZ3A2CnJ3._2VoB-BqDjcx41XiO-Sq6iJ')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(3) > div._1OwJclDo5JkkZuvbXf3bOZ > div._3BshymDCc0OSBa0oS0zJj- > div > div > div > div._2taOCr9V0CnonoT0H_eksA > span:nth-child(2)')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_id('searchName').send_keys(id)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(3) > div._1OwJclDo5JkkZuvbXf3bOZ > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(3) > div._19XtTBEROGh2PcaaoNIpB_ > div._1ICwZfGa-u0vwVoGlQwVkD > table > tbody > tr > td:nth-child(2)')
    btn_elm.click()
    sleep(1)
    # 삭제
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div._1ScRXCyfLzSgOrfscbk2CO > div._1kbII_WxXJa1Ewh8pE8MFE > button:nth-child(3)')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[3]/button[1]')
    btn_elm.click()
    sleep(1)
    # 고객 등록 및 조회
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div.ICjgBEGgF_4nnT6DOdQ9t > div > dl:nth-child(1) > dd:nth-child(2) > a > button')
    btn_elm.click()
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._1OwJclDo5JkkZuvbXf3bOZ > div._3BshymDCc0OSBa0oS0zJj- > div > div > div')
    btn_elm.click()
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._1OwJclDo5JkkZuvbXf3bOZ > div._3BshymDCc0OSBa0oS0zJj- > div > div > div > div._2taOCr9V0CnonoT0H_eksA > span:nth-child(2)')
    btn_elm.click()
    sleep(1)
    elm = driver.find_element_by_id('searchName').send_keys(id)
    sleep(1)
    btn_elm = driver.find_element_by_css_selector('#root > div > div._3QplAkdvLv6royGAnjzvaE > div._2C_pOoOejXSb-D5ebFF68J > div > div:nth-child(2) > div > div > div._1OwJclDo5JkkZuvbXf3bOZ > button')
    btn_elm.click()
    sleep(1)
    try:
        elm = driver.find_element_by_xpath('/html/body/div/div/div[2]/div[2]/div/div[2]/div/div/div[2]/div[2]/table/tbody/tr/td[contains(text(), "없습니다.")]')
        ws['B32'] = '13'
        ws['C32'] = 'FAIL'
        return 0
    except:
        ws['B32'] = '13'
        ws['C32'] = 'PASS'
        return 1

def send(file):
    sendEmail = "yoojung_jang@tmax.co.kr"
    recvEmail = "houngki_park@tmax.co.kr"
    ccEmail = ["yujeong_boo@tmax.co.kr", "seungeun_kim@tmax.co.kr", "jia_park@tmax.co.kr"]
    password = "apdlfdla7!"

    smtpName = "mail.tmax.co.kr"
    smtpPort = 587

    msg = MIMEMultipart()

    msg['Subject'] = "Admin 자동화 테스트 결과 공유"
    msg['From'] = sendEmail
    msg['To'] = recvEmail
    msg['Cc'] = ",".join(ccEmail)

    # 본문 추가
    text = "안녕하세요.\n\nAdmin 자동화 테스트 결과 공유드립니다.\n\n테스트 케이스 별 자세한 항목 정리 및 결과는 아래 시트에 적어놓았습니다.\n" \
           "https://docs.google.com/spreadsheets/d/1GPk076PkLIo-F2hED8hddjXb8_jn0d_gnuQKv-s5GX0/edit#gid=813980518\n\n" \
           "감사합니다.\n장유정 드림."
    contentPart = MIMEText(text)
    msg.attach(contentPart)

    # 파일 추가
    Filename = file
    with open(Filename, 'rb') as etcFD:
        etcPart = MIMEApplication(etcFD.read())
        etcPart.add_header('Content-Disposition', 'attachment', filename=Filename)
        msg.attach(etcPart)

    s = smtplib.SMTP(smtpName, smtpPort)
    s.starttls()
    s.login(sendEmail, password)
    recvEmail = ["yoojung_jang@tmax.co.kr", "houngki_park@tmax.co.kr", "yujeong_boo@tmax.co.kr", "seungeun_kim@tmax.co.kr", "jia_park@tmax.co.kr"]
    s.sendmail(sendEmail, recvEmail, msg.as_string())
    s.close()



# 메인 시작
if __name__ == '__main__':

    # id2는 파일에서도 쓰이고, id4는 파일에서만 쓰임 (신규 계정들)
    count = input("오늘 시작할 id 번호는? ")
    num_count = int(count)
    id = 'admintest' + count + '@tmax.co.kr'
    id2 = 'admintest' + str(num_count+1) + '@tmax.co.kr'
    id3 = 'admintest' + str(num_count+2) + '@tmax.co.kr'
    testid = 'admintest0@tmax.co.kr'

    pass_count = 0

    # 테스트 결과 남길 엑셀파일 및 시트 생성
    wb = Workbook()
    wb.remove(wb['Sheet'])
    ws = wb.create_sheet('Result')

    ws = wb.active
    now = datetime.now()
    ws['B2'] = 'Test Start Time: ' + str(now)
    ws['B4'] = 'TC_Number'
    ws['C4'] = 'Result'

    # 카메라, 마이크 허용
    opt = Options()
    opt.add_argument("--disable-infobars")
    opt.add_argument("start-maximized")
    opt.add_argument("--disable-extensions")

    opt.add_experimental_option("prefs", { \
        "profile.default_content_setting_values.media_stream_camera": 1,
        "profile.default_content_setting_values.media_stream_mic": 1})

    # 시작
    driver = webdriver.Chrome(options=opt, executable_path=r'C:\chromedriver.exe')

    driver.implicitly_wait(15)

    driver.get('https://stage.hypermeeting.biz')
    sleep(5)

    pass_count += test29()
    pass_count += test11()
    pass_count += test10()
    pass_count += test1()
    pass_count += test2(id)
    pass_count += test435()
    pass_count += test7()
    pass_count += test8(id)
    pass_count += test9(id2)
    pass_count += test614()
    pass_count += test24()
    pass_count += test16()
    pass_count += test15()
    pass_count += test17(id3)
    pass_count += test201921()
    pass_count += test25(id)
    pass_count += test27()
    pass_count += test1826()
    pass_count += test22()
    pass_count += test28(id2)
    pass_count += test23()
    pass_count += test12()
    pass_count += test13(id2)

    ws['B34'] = 'Result-PASS: ' + str(pass_count) + '/28'
    ws['B35'] = 'Result-FAIL: ' + str(28-pass_count) + '/28'
    now = datetime.now()
    ws['B36'] = 'Test End Time: ' + str(now)

    # 저장 전 너비
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

    wb.save('C:\\admin\\test_result.xlsx')

    driver.quit()

    # 결과 파일 메일로 전송 - 주의!
    # send('C:\\admin\\test_result.xlsx')
